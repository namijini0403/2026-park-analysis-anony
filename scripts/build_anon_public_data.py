from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKSPACE = Path(os.environ.get("SOURCE_WORKSPACE", ROOT.parent)).resolve()
ANON_XLSX = Path(os.environ.get("ANON_CODE_XLSX", WORKSPACE / "incheon_school_anon_codes.xlsx")).resolve()
DATA_DIR = Path(os.environ.get("SOURCE_DATA_PROCESSED", WORKSPACE / "2026-park-analysis" / "data_processed")).resolve()
OUT_DIR = ROOT / "public" / "data"


LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
EARTH_RADIUS_M = 6_371_000


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
        f.write("\n")


def n(value: Any, fallback: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return fallback
        result = float(value)
        return result if math.isfinite(result) else fallback
    except (TypeError, ValueError):
        return fallback


def rounded(value: Any, digits: int = 1) -> float:
    return round(n(value), digits)


def safe_int(value: Any) -> int:
    return int(round(n(value)))


def school_id(row: dict[str, Any]) -> str:
    return str(row.get("학교ID") or row.get("school_id") or "").strip()


def school_name(row: dict[str, Any]) -> str:
    return str(row.get("학교명") or row.get("school_name") or "").strip()


def distance_m(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    dlat = math.radians(lat2 - lat1)
    dlng = math.radians(lng2 - lng1)
    a = (
        math.sin(dlat / 2) ** 2
        + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlng / 2) ** 2
    )
    return 2 * EARTH_RADIUS_M * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def local_xy_m(origin_lat: float, origin_lng: float, lat: float, lng: float) -> tuple[float, float]:
    x = math.radians(lng - origin_lng) * EARTH_RADIUS_M * math.cos(math.radians(origin_lat))
    y = math.radians(lat - origin_lat) * EARTH_RADIUS_M
    return x, y


def rotate_and_flip(x: float, y: float, anon_code: str) -> tuple[float, float]:
    digest = hashlib.sha256(anon_code.encode("utf-8")).digest()
    angle = (int.from_bytes(digest[:2], "big") % 360) * math.pi / 180
    flip = -1 if digest[2] % 2 else 1
    xr = x * math.cos(angle) - y * math.sin(angle)
    yr = x * math.sin(angle) + y * math.cos(angle)
    return round(xr * flip, 1), round(yr, 1)


def deterministic_unit(anon_code: str, salt: str) -> float:
    digest = hashlib.sha256(f"{anon_code}:{salt}".encode("utf-8")).digest()
    return int.from_bytes(digest[:4], "big") / 0xFFFFFFFF


def synthetic_walk_shape(anon_code: str) -> list[dict[str, float]]:
    points = []
    for index in range(12):
        angle = (math.pi * 2 * index) / 12
        factor = 0.72 + deterministic_unit(anon_code, f"walk-{index}") * 0.28
        radius = 500 * factor
        points.append(
            {
                "x_m": round(math.cos(angle) * radius, 1),
                "y_m": round(math.sin(angle) * radius, 1),
            }
        )
    return points


def synthetic_barriers(row: dict[str, Any], anon_code: str) -> list[dict[str, Any]]:
    count = max(
        safe_int(row.get("nearest_official_major_road_crossing_count")),
        safe_int(row.get("nearest_functional_major_road_crossing_count")),
    )
    if count <= 0:
        return []
    barriers = []
    for index in range(min(count, 3)):
        angle = deterministic_unit(anon_code, f"barrier-angle-{index}") * math.pi * 2
        offset = 130 + deterministic_unit(anon_code, f"barrier-offset-{index}") * 260
        length = 160 + deterministic_unit(anon_code, f"barrier-length-{index}") * 160
        cx = math.cos(angle) * offset
        cy = math.sin(angle) * offset
        dx = math.cos(angle + math.pi / 2) * length / 2
        dy = math.sin(angle + math.pi / 2) * length / 2
        barriers.append(
            {
                "label": anon_label(index, "단절요소"),
                "kind": "road_barrier",
                "x1_m": round(cx - dx, 1),
                "y1_m": round(cy - dy, 1),
                "x2_m": round(cx + dx, 1),
                "y2_m": round(cy + dy, 1),
                "crossing_count": index + 1,
            }
        )
    return barriers


def global_relative_index_positions(rows: list[dict[str, Any]]) -> dict[str, dict[str, float]]:
    valid_rows = [
        row
        for row in rows
        if math.isfinite(row["lat"]) and math.isfinite(row["lng"])
    ]
    if not valid_rows:
        return {}

    center_lat = sum(row["lat"] for row in valid_rows) / len(valid_rows)
    center_lng = sum(row["lng"] for row in valid_rows) / len(valid_rows)
    transformed = []
    for row in valid_rows:
        x, y = local_xy_m(center_lat, center_lng, row["lat"], row["lng"])
        mask_angle = deterministic_unit(row["anon_code"], "public-map-mask-angle") * math.pi * 2
        mask_distance = 650 + deterministic_unit(row["anon_code"], "public-map-mask-distance") * 600
        masked_x = x + math.cos(mask_angle) * mask_distance
        masked_y = y + math.sin(mask_angle) * mask_distance
        transformed.append((row["anon_code"], masked_x, masked_y))

    min_x = min(x for _, x, _ in transformed)
    max_x = max(x for _, x, _ in transformed)
    min_y = min(y for _, _, y in transformed)
    max_y = max(y for _, _, y in transformed)
    mid_x = (min_x + max_x) / 2
    mid_y = (min_y + max_y) / 2
    scale = max((max_x - min_x) / 700, (max_y - min_y) / 460, 1.0)
    positions: dict[str, dict[str, float]] = {}
    for anon_code, x, y in transformed:
        positions[anon_code] = {
            "x": round(min(760, max(40, 400 + (x - mid_x) / scale)), 1),
            "y": round(min(500, max(40, 270 - (y - mid_y) / scale)), 1),
        }
    return positions


def anon_label(index: int, prefix: str) -> str:
    if index < len(LETTERS):
        return f"{prefix} {LETTERS[index]}"
    return f"{prefix} {LETTERS[index % len(LETTERS)]}{index // len(LETTERS) + 1}"


def load_anon_mapping() -> tuple[dict[str, dict[str, str]], dict[str, dict[str, str]]]:
    wb = load_workbook(ANON_XLSX, data_only=True, read_only=True)
    ws = wb["내부용_실명매핑"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(v) for v in rows[0]]
    by_id: dict[str, dict[str, str]] = {}
    by_name: dict[str, dict[str, str]] = {}
    for row_values in rows[1:]:
        row = {headers[i]: "" if row_values[i] is None else str(row_values[i]) for i in range(len(headers))}
        item = {
            "anon_code": row["anon_code"],
            "display_name": row["anon_display_name"],
            "short_label": row["anon_short_label"],
            "gu": row.get("gu", ""),
        }
        by_id[row["학교ID"]] = item
        by_name[row["학교명"]] = item
    return by_id, by_name


def row_map(rows: list[dict[str, str]], key_fn) -> dict[str, dict[str, str]]:
    out = {}
    for row in rows:
        key = key_fn(row)
        if key:
            out[key] = row
    return out


def group_trend(rows: list[dict[str, str]]) -> dict[str, list[dict[str, int]]]:
    grouped: dict[str, list[dict[str, int]]] = defaultdict(list)
    for row in rows:
        sid = school_id(row)
        if not sid:
            continue
        grouped[sid].append(
            {
                "year": safe_int(row.get("연도") or row.get("year")),
                "students": safe_int(row.get("학생수") or row.get("students")),
            }
        )
    for values in grouped.values():
        values.sort(key=lambda item: item["year"])
    return grouped


def parse_linked_names(value: Any) -> list[str]:
    text = str(value or "")
    if not text:
        return []
    return sorted(set(match.strip() for match in re.findall(r"'([^']+)'", text) if match.strip()))


def case_policy(case_type: Any) -> str:
    key = safe_int(case_type)
    return {
        1: "즉시 개선 대상",
        2: "우선 검토 대상",
        3: "모니터링 대상",
        4: "유지·관리 대상",
        99: "별도 정책 필요",
    }.get(key, "검토 대상")


def normalized_case_type(row: dict[str, Any]) -> int:
    if safe_int(row.get("is_separate_bundle_tag")) == 1 or safe_int(row.get("is_island_tag")) == 1:
        return 99
    value = safe_int(row.get("case_type"))
    return value if value in {1, 2, 3, 4, 99} else 99


def load_candidate_features(by_name: dict[str, dict[str, str]]) -> dict[str, list[dict[str, Any]]]:
    geojson_path = DATA_DIR / "candidate_grid_final.geojson"
    with geojson_path.open("r", encoding="utf-8") as f:
        geojson = json.load(f)

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for feature in geojson.get("features", []):
        props = feature.get("properties") or {}
        for linked_name in parse_linked_names(props.get("linked_schools")):
            mapping = by_name.get(linked_name)
            if not mapping:
                continue
            grouped[mapping["anon_code"]].append(dict(props))

    for key in list(grouped):
        grouped[key].sort(
            key=lambda item: (
                n(item.get("candidate_rank_mixed"), 999999),
                -n(item.get("walkshed_beneficiary_2029")),
            )
        )
        grouped[key] = grouped[key][:8]
    return grouped


def nearest_points(
    school_lat: float,
    school_lng: float,
    rows: list[dict[str, str]],
    lat_key: str,
    lng_key: str,
    radius_m: float,
    limit: int,
) -> list[tuple[dict[str, str], float]]:
    points = []
    for row in rows:
        lat = n(row.get(lat_key), float("nan"))
        lng = n(row.get(lng_key), float("nan"))
        if not math.isfinite(lat) or not math.isfinite(lng):
            continue
        dist = distance_m(school_lat, school_lng, lat, lng)
        if dist <= radius_m:
            points.append((row, dist))
    points.sort(key=lambda item: item[1])
    return points[:limit]


def synthetic_point(
    anon_code: str,
    origin_lat: float,
    origin_lng: float,
    lat: float,
    lng: float,
    label: str,
    kind: str,
    distance: float,
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    x, y = local_xy_m(origin_lat, origin_lng, lat, lng)
    xr, yr = rotate_and_flip(x, y, anon_code)
    return {
        "label": label,
        "kind": kind,
        "x_m": xr,
        "y_m": yr,
        "distance_m": safe_int(distance),
        **(extra or {}),
    }


def route_meta(candidate_routes: dict[str, Any], sid: str, grid_id: str) -> dict[str, Any]:
    item = (candidate_routes.get(sid) or {}).get(grid_id) or {}
    counts = item.get("counts") or {}
    return {
        "route_length_m": rounded(item.get("route_length_m"), 1),
        "barrier_label": str(item.get("severity_label") or "경로 정보 없음"),
        "barrier_counts": {
            "primary": safe_int(counts.get("primary")),
            "secondary": safe_int(counts.get("secondary")),
            "tertiary": safe_int(counts.get("tertiary")),
        },
    }


def build_school_payloads() -> tuple[list[dict[str, Any]], dict[str, Any], dict[str, dict[str, float]]]:
    by_id, by_name = load_anon_mapping()
    priority_rows = read_csv(DATA_DIR / "school_priority_with_functional_park_layer.csv")
    coord_rows = read_csv(DATA_DIR / "schools.csv")
    student_rows = read_csv(DATA_DIR / "student_trend.csv")
    forecast_rows = read_csv(DATA_DIR / "school_enrollment_forecast_20260418_model1.csv")
    similar_rows = read_csv(DATA_DIR / "school_similar_schools_top5.csv")
    parks = read_csv(DATA_DIR / "parks.csv")
    redevelopment = read_csv(DATA_DIR / "redevelopment_geocoded.csv")
    apartments = read_csv(DATA_DIR / "large_apt_complexes_2025.csv")
    with (DATA_DIR / "candidate_barrier_routes_by_school.json").open("r", encoding="utf-8") as f:
        candidate_routes = json.load(f)

    coords_by_id = row_map(coord_rows, school_id)
    forecast_by_id = row_map(forecast_rows, school_id)
    similar_by_id = row_map(similar_rows, school_id)
    trends_by_id = group_trend(student_rows)
    candidates_by_school = load_candidate_features(by_name)

    schools: list[dict[str, Any]] = []
    global_coord_rows: list[dict[str, Any]] = []
    for priority in priority_rows:
        sid = school_id(priority)
        mapping = by_id.get(sid)
        coords = coords_by_id.get(sid, {})
        if not mapping or not coords:
            continue
        anon_code = mapping["anon_code"]
        lat = n(coords.get("위도"), float("nan"))
        lng = n(coords.get("경도"), float("nan"))
        if not math.isfinite(lat) or not math.isfinite(lng):
            continue
        global_coord_rows.append({"anon_code": anon_code, "lat": lat, "lng": lng})

        forecast = forecast_by_id.get(sid, {})
        similar = similar_by_id.get(sid, {})
        nearby_parks = nearest_points(lat, lng, parks, "위도", "경도", 900, 3)
        nearby_redev = nearest_points(lat, lng, redevelopment, "위도", "경도", 650, 3)
        nearby_apts = [
            (row, dist)
            for row, dist in nearest_points(lat, lng, apartments, "위도", "경도", 650, 3)
            if n(row.get("세대수")) >= 500
        ]

        map_points: list[dict[str, Any]] = [
            {"label": "학교", "kind": "school", "x_m": 0, "y_m": 0, "distance_m": 0}
        ]
        for index, (row, dist) in enumerate(nearby_parks):
            map_points.append(
                synthetic_point(anon_code, lat, lng, n(row.get("위도")), n(row.get("경도")), anon_label(index, "공원"), "park", dist)
            )
        for index, (row, dist) in enumerate(nearby_apts):
            households = safe_int(row.get("세대수"))
            map_points.append(
                synthetic_point(
                    anon_code,
                    lat,
                    lng,
                    n(row.get("위도")),
                    n(row.get("경도")),
                    anon_label(index, "단지"),
                    "apartment",
                    dist,
                    {"households_bucket": f"{max(500, households // 100 * 100)}세대 이상"},
                )
            )
        for index, (row, dist) in enumerate(nearby_redev):
            map_points.append(
                synthetic_point(
                    anon_code,
                    lat,
                    lng,
                    n(row.get("위도")),
                    n(row.get("경도")),
                    anon_label(index, "구역"),
                    "redevelopment",
                    dist,
                    {"stage": str(row.get("진행단계") or "단계 정보 없음")},
                )
            )

        school_candidates = []
        for index, candidate in enumerate(candidates_by_school.get(anon_code, [])[:6]):
            cx = n(candidate.get("cx"))
            cy = n(candidate.get("cy"))
            label = anon_label(index, "후보지")
            meta = route_meta(candidate_routes, sid, str(candidate.get("grid_id") or ""))
            point = synthetic_point(anon_code, lat, lng, cy, cx, label, "candidate", distance_m(lat, lng, cy, cx))
            map_points.append(point)
            school_candidates.append(
                {
                    "label": label,
                    "walkshed_beneficiary_2029": safe_int(candidate.get("walkshed_beneficiary_2029")),
                    "walkshed_beneficiary_2031": safe_int(candidate.get("walkshed_beneficiary_2031")),
                    "nearest_park_dist_m": rounded(candidate.get("nearest_park_dist"), 1),
                    "nearest_playground_dist_m": rounded(candidate.get("nearest_pg_dist"), 1),
                    "land_feasibility_level": str(candidate.get("land_feasibility_level") or "medium"),
                    "redev_flag": str(candidate.get("redev_flag")).lower() == "true",
                    **meta,
                }
            )

        similar_schools = []
        for rank in range(1, 6):
            peer_name = str(similar.get(f"similar_school_{rank}_name") or "")
            peer_mapping = by_name.get(peer_name)
            if not peer_mapping:
                continue
            similar_schools.append(
                {
                    "rank": rank,
                    "anon_code": peer_mapping["anon_code"],
                    "display_name": peer_mapping["display_name"],
                    "gu": str(similar.get(f"similar_school_{rank}_gu") or ""),
                    "nearest_park_dist_m": rounded(similar.get(f"similar_school_{rank}_nearest_park_dist_m"), 1),
                    "green_ratio": rounded(similar.get(f"similar_school_{rank}_iso_green_ratio"), 2),
                    "playground_count": safe_int(similar.get(f"similar_school_{rank}_iso_playground_count")),
                }
            )

        trends = trends_by_id.get(sid, [])
        current_students = trends[-1]["students"] if trends else safe_int(similar.get("current_students_2025"))
        school_payload = {
            "anon_code": anon_code,
            "display_name": mapping["display_name"],
            "short_label": mapping["short_label"],
            "gu": str(priority.get("gu") or mapping.get("gu") or ""),
            "case_type": normalized_case_type(priority),
            "case_policy_label": case_policy(normalized_case_type(priority)),
            "case_status_label": str(priority.get("access_condition_label") or priority.get("case_label") or "검토 대상"),
            "priority_rank": safe_int(priority.get("priority_rank")),
            "metrics": {
                "nearest_park_dist_m": rounded(priority.get("nearest_park_dist_m"), 1),
                "official_park_count_500m": safe_int(priority.get("iso_official_park_count") or priority.get("iso_park_count")),
                "functional_park_count_500m": safe_int(priority.get("iso_functional_park_count")),
                "playground_count_500m": safe_int(priority.get("iso_playground_count")),
                "green_ratio": rounded(priority.get("display_green_ratio") or priority.get("corrected_green_ratio") or priority.get("iso_green_ratio"), 2),
                "current_students_2025": current_students,
                "potential_demand_2029": safe_int(forecast.get("forecast_2029") or forecast.get("predicted_2029") or priority.get("forecast_2029")),
                "potential_demand_2031": safe_int(forecast.get("forecast_2031") or forecast.get("predicted_2031") or priority.get("forecast_2031")),
            },
            "trend": trends,
            "similar_schools": similar_schools,
            "candidates": school_candidates,
            "synthetic_map": {
                "note": "실제 좌표가 아닌 학교 중심 상대거리 도식입니다. 학교별 고정 난수 회전/반전을 적용했습니다.",
                "radius_500m": 500,
                "walk_500m_shape": synthetic_walk_shape(anon_code),
                "barriers": synthetic_barriers(priority, anon_code),
                "points": map_points,
            },
        }
        schools.append(school_payload)

    schools.sort(key=lambda item: (item["gu"], item["short_label"]))
    map_positions = global_relative_index_positions(global_coord_rows)

    by_gu: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in schools:
        by_gu[item["gu"]].append(item)

    statistics = {
        "summary": {
            "school_count": len(schools),
            "district_count": len(by_gu),
            "urgent_count": sum(1 for item in schools if item["case_type"] == 1),
            "priority_count": sum(1 for item in schools if item["case_type"] == 2),
            "potential_demand_2029": sum(item["metrics"]["potential_demand_2029"] for item in schools),
        },
        "districts": [
            {
                "gu": gu,
                "school_count": len(items),
                "urgent_count": sum(1 for item in items if item["case_type"] == 1),
                "priority_count": sum(1 for item in items if item["case_type"] == 2),
                "avg_nearest_park_dist_m": rounded(sum(item["metrics"]["nearest_park_dist_m"] for item in items) / max(len(items), 1), 1),
                "avg_green_ratio": rounded(sum(item["metrics"]["green_ratio"] for item in items) / max(len(items), 1), 2),
                "top_priority": [
                    {
                        "anon_code": item["anon_code"],
                        "display_name": item["display_name"],
                        "case_policy_label": item["case_policy_label"],
                        "nearest_park_dist_m": item["metrics"]["nearest_park_dist_m"],
                        "potential_demand_2029": item["metrics"]["potential_demand_2029"],
                    }
                    for item in sorted(items, key=lambda row: row["priority_rank"] or 999999)[:5]
                ],
            }
            for gu, items in sorted(by_gu.items())
        ],
    }

    return schools, statistics, map_positions


def build_split_payloads(
    schools: list[dict[str, Any]],
    statistics: dict[str, Any],
    map_positions: dict[str, dict[str, float]],
) -> dict[str, Any]:
    map_index = {
        "schools": [
            {
                "anon_code": school["anon_code"],
                "display_name": school["display_name"],
                "short_label": school["short_label"],
                "gu": school["gu"],
                "case_type": school["case_type"],
                "case_policy_label": school["case_policy_label"],
                **map_positions.get(school["anon_code"], {"x": 400.0, "y": 270.0}),
            }
            for school in schools
        ]
    }

    school_detail = {
        "schools": [
            {
                "anon_code": school["anon_code"],
                "display_name": school["display_name"],
                "short_label": school["short_label"],
                "gu": school["gu"],
                "case_type": school["case_type"],
                "case_policy_label": school["case_policy_label"],
                "case_status_label": school["case_status_label"],
                "priority_rank": school["priority_rank"],
                "metrics": school["metrics"],
                "trend": school["trend"],
                "similar_schools": school["similar_schools"],
                "interpretation": {
                    "current_gap": "공원 접근, 녹지 비율, 놀이터 수를 분리해 검토합니다.",
                    "future_demand": "현재 학생 수와 2029/2031 잠재 수요를 함께 봅니다.",
                    "access_friction": "보행 부담과 단절요소는 후보지 판단의 보조 신호입니다.",
                },
            }
            for school in schools
        ]
    }

    candidate_anon = {
        "schools": [
            {
                "anon_code": school["anon_code"],
                "display_name": school["display_name"],
                "candidates": school["candidates"],
            }
            for school in schools
        ]
    }

    synthetic_map = {
        "schools": [
            {
                "anon_code": school["anon_code"],
                "display_name": school["display_name"],
                "gu": school["gu"],
                "case_type": school["case_type"],
                "synthetic_map": school["synthetic_map"],
            }
            for school in schools
        ]
    }

    app_summary = {
        "title": "반경 너머, 도달 가능성으로",
        "subtitle": "비식별 학교 단위 야외활동 환경 정책지원 시스템",
        "principles": [
            "학교명과 실제 좌표를 공개하지 않습니다.",
            "현재 격차, 미래 수요, 접근 마찰, 후보지 추천을 분리해 설명합니다.",
            "AI는 자동 결정이 아니라 조정 가능한 보조 신호로 사용합니다.",
        ],
        "flow": [
            "첫 화면",
            "비식별 도식지도",
            "학교 상세 리포트",
            "후보지 시뮬레이션",
            "전체 통계",
        ],
        "statistics_summary": statistics["summary"],
    }

    return {
        "map_index_anon.json": map_index,
        "school_detail_anon.json": school_detail,
        "candidate_anon.json": candidate_anon,
        "synthetic_map_anon.json": synthetic_map,
        "app_summary_anon.json": app_summary,
    }


def main() -> None:
    for path in [ANON_XLSX, DATA_DIR / "school_priority_with_functional_park_layer.csv"]:
        if not path.exists():
            raise SystemExit(f"Missing local input: {path}")

    schools, statistics, map_positions = build_school_payloads()
    generated_at = datetime.now(timezone.utc).isoformat(timespec="seconds")

    write_json(OUT_DIR / "schools_anon.json", {"generated_at": generated_at, "schools": schools})
    write_json(OUT_DIR / "statistics_anon.json", {"generated_at": generated_at, **statistics})
    split_payloads = build_split_payloads(schools, statistics, map_positions)
    for file_name, payload in split_payloads.items():
        write_json(OUT_DIR / file_name, {"generated_at": generated_at, **payload})
    files = ["schools_anon.json", "statistics_anon.json", *sorted(split_payloads)]
    write_json(
        OUT_DIR / "manifest.json",
        {
            "generated_at": generated_at,
            "source_policy": "Local private source data was transformed into anonymized JSON only.",
            "files": files,
        },
    )
    print(f"Wrote {len(schools)} anonymized schools to {OUT_DIR}")


if __name__ == "__main__":
    main()
